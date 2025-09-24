import requests
import json
from datetime import datetime, timedelta, time as dt_time
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
import time

# ================== 配置 ==================
API_TOKEN = "BEDDU1T7F5R7D9U9Z0V8I7F5"
URL_BIND_IP = "https://api.csqaq.com/api/v1/sys/bind_local_ip"
URL_CHART = "https://api.csqaq.com/api/v1/info/chart"


#输入商品的时候请
goods = [
    {"good_id": "19653", "name": "传承-崭新","category":"gun","item":"传承"},
    #{"good_id": "19619", "name": "传承-略磨","category":"gun","item":"传承"},
    #{"good_id": "144", "name": "皇后","category":"gun","item":"皇后"},
    {"good_id": "14797", "name": "夜愿","category":"gun","item":"夜愿"},
    {"good_id": "16422", "name": "一发入魂","category":"gun","item":"一发入魂"},
    {"good_id": "23718", "name": "幽独","category":"gun","item":"幽独"},
    {"good_id": "12196", "name": "deagle蓝色层压板","category":"gun","item":"deagle蓝色层压板"},
    {"good_id": "20607", "name": "deagle跷跷板","category":"gun","item":"deagle跷跷板"},
    {"good_id": "20565", "name": "鼠鼠我呀","category":"gun","item":"鼠鼠我呀"},
    {"good_id": "1205", "name": "血虎","category":"gun","item":"血虎"},
    {"good_id": "14135", "name": "黑蛋-全息","category":"stick","item":"黑蛋"},
]

OUTPUT_FILE = "all_goods_wide.xlsx"
FIXED_TIMES = [dt_time(13, 0), dt_time(20, 0), dt_time(23, 59)]  # 每日抓取固定时间

# ================== 绑定 IP ==================
def bind_my_ip():
    headers = {"ApiToken": API_TOKEN}
    try:
        resp = requests.post(URL_BIND_IP, headers=headers, timeout=10)
        if resp.status_code == 200:
            j = resp.json()
            if j.get("code") == 200:
                print("✅ IP 绑定成功：", j.get("data"))
                return True
            else:
                print("❌ 绑定失败:", j)
        else:
            print("❌ HTTP 错误:", resp.status_code, resp.text)
    except Exception as e:
        print("❌ 请求绑定IP失败:", e)
    return False

# ================== 数据抓取（增加字段） ==================
def fetch_price_history_extended(good_id, period=90, retries=3, delay=2):
    headers = {"ApiToken": API_TOKEN, "Content-Type": "application/json"}
    payload = {
        "good_id": good_id,
        "key": "sell_price",
        "platform": 1,
        "period": str(period),
        "style": "all_style"
    }
    for attempt in range(1, retries+1):
        try:
            response = requests.post(URL_CHART, headers=headers, data=json.dumps(payload), timeout=10)
            if response.status_code == 200:
                res_json = response.json()
                if res_json.get("code") == 200:
                    data = res_json.get("data", {})
                    timestamps = data.get("timestamp", [])
                    prices = data.get("main_data", [])
                    volumes = data.get("num_data", [])
                    max_prices = data.get("max_price", [])
                    min_prices = data.get("min_price", [])
                    if timestamps and prices:
                        df = pd.DataFrame({
                            "timestamp": [datetime.fromtimestamp(ts/1000) for ts in timestamps],
                            "price": prices,
                            "volume": volumes if volumes else [0]*len(prices),
                            "max_price": pd.Series(prices).cummax() if not max_prices else max_prices,
                            "min_price": pd.Series(prices).cummin() if not min_prices else min_prices,
                            "price_change": [prices[i] - prices[i-1] if i > 0 else 0 for i in range(len(prices))]
                        })
                        df['price_change'] = df['price'].diff().fillna(0)  # 价格差
                        return df
            print(f"[HTTP {response.status_code}] good_id={good_id} 返回异常或为空")
        except Exception as e:
            print(f"⚠ 尝试 {attempt}/{retries} 获取 good_id={good_id} 失败: {e}")
        time.sleep(delay)
    print(f"❌ good_id={good_id} 所有尝试失败")
    return pd.DataFrame()

# ================== Excel  ==================（之后修改列表的列名，使得后续数据处理时，读取数据正常，price_change部分仍有相当大的问题）
def update_wide_excel(goods_data_dict):
    """
    goods_data_dict: { '饰品名': df, ... }
    goods: 全局列表，需包含 'category' 和 'item' 字段
    """

    # 构建类别 -> item -> 饰品 的映射
    category_map = {}
    for good in goods:
        cat = good["category"]
        item = good["item"]
        category_map.setdefault(cat, {}).setdefault(item, []).append(good["name"])

    for category, item_dict in category_map.items():
        output_file = f"{category}.xlsx"  # 每个类别生成一个 Excel，名字设置为category
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
        else:
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            #wb.active.title = "wide"  # 默认 sheet

        for item, good_names in item_dict.items():  #这部分需要修改，将每个sheet的列名修改成通用的列名，便于后续处理
            # 每个 item 放一个 sheet
            if item in wb.sheetnames:
                ws = wb[item]
            else:
                ws = wb.create_sheet(item)

            all_rows = []
            for name in good_names:
                df = goods_data_dict.get(name)
                if df is None or df.empty:
                    continue

                # 整理数据按固定时间抓取
                df['date'] = df['timestamp'].dt.date
                grouped = df.groupby('date')
                rows = []
                for day, group in grouped:
                    group = group.sort_values('timestamp').reset_index(drop=True)
                    for ft in FIXED_TIMES:
                        target_dt = datetime.combine(day, ft)
                        now = datetime.now()
                        if target_dt > now:
                            continue
                        closest_row = group.iloc[(group['timestamp'] - target_dt).abs().argsort().iloc[0]]
                        row = {
                            "timestamp": target_dt,
                            "price": closest_row['price'],
                            "volume": closest_row['volume'],
                            "max_price": closest_row.get('max_price', 0),
                            "min_price": closest_row.get('min_price', 0)
                        }
                        rows.append(row)

                if rows:
                    all_rows.append(pd.DataFrame(rows))

            if not all_rows:
                continue

            # 合并所有饰品数据
            combined_df = pd.DataFrame({"timestamp": sorted({row['timestamp'] for df in all_rows for _, row in df.iterrows()})})
            for df in all_rows:
                combined_df = pd.merge(combined_df, df, on="timestamp", how="left")
            combined_df.sort_values("timestamp", inplace=True)

            # 写表头
            if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
                for col_idx, col_name in enumerate(combined_df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)

            # 增量更新
            existing_rows = ws.max_row
            last_ts = ws.cell(existing_rows, 1).value if existing_rows > 1 else None
            new_data = combined_df
            if last_ts:
                new_data = combined_df[combined_df["timestamp"] > pd.to_datetime(last_ts)]

            for r_idx, row in enumerate(new_data.itertuples(index=False), existing_rows + 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file)
        print(f"{output_file} 已更新完成")



# ================== 主执行 ==================
def run_update():
    if not bind_my_ip():
        print("❌ IP绑定失败，程序退出")
        return
    print(f"--- 数据刷新 {datetime.now()} ---")
    goods_data = {}
    for good in goods:
        df = fetch_price_history_extended(good["good_id"])
        goods_data[good["name"]] = df
        if not df.empty:
            print(f"{good['name']} 最新价格: {df['price'].iloc[-1]:.2f} RMB")
    update_wide_excel(goods_data)
    print("程序执行完毕")

if __name__ == "__main__":
    run_update()
