import requests
import json
from datetime import datetime, timedelta, time as dt_time
import pandas as pd
import os
from openpyxl import load_workbook, Workbook

# ================== 配置 ==================
API_TOKEN = "BEDDU1T7F5R7D9U9Z0V8I7F5"
url = "https://api.csqaq.com/api/v1/info/chart"

goods = [
    {"good_id": "19653", "name": "传承"},
    {"good_id": "144", "name": "皇后"},
    {"good_id": "14797", "name": "夜愿"},
    {"good_id": "16422", "name": "一发入魂"},
    {"good_id": "23718", "name": "幽独"},
]

OUTPUT_FILE = "all_goods_wide.xlsx"
FIXED_TIMES = [dt_time(10, 0), dt_time(22, 0)]  # 每日固定抓取时间


# ================== 数据抓取 ==================
def fetch_price_history(good_id):
    payload = json.dumps({
        "good_id": good_id,
        "key": "sell_price",
        "platform": 1,
        "period": "90",  # 最近90天
        "style": "all_style"
    })
    headers = {
        'ApiToken': API_TOKEN,
        'Content-Type': 'application/json'
    }

    try:
        response = requests.post(url, headers=headers, data=payload)
        if response.status_code == 200:
            res_json = response.json()
            if res_json.get('code') == 200:
                data = res_json.get('data', {})
                timestamps = data.get('timestamp', [])
                prices = data.get('main_data', [])
                volumes = data.get('num_data', [])
                if timestamps and prices:
                    df = pd.DataFrame({
                        "timestamp": [datetime.fromtimestamp(ts / 1000) for ts in timestamps],
                        "price": prices,
                        "volume": volumes if volumes else [0]*len(prices)
                    })
                    # 计算相对前一天价格的绝对变化量
                    df['abs_change'] = df['price'].diff().fillna(0)
                    return df
        print(f"请求饰品 {good_id} 返回数据异常或为空")
        return pd.DataFrame()
    except Exception as e:
        print(f"获取饰品 {good_id} 数据失败: {e}")
        return pd.DataFrame()


# ================== Excel 更新（保留格式） ==================
def update_wide_excel(goods_data_dict):
    all_rows = []
    for name, df in goods_data_dict.items():
        if df.empty:
            continue
        df['date'] = df['timestamp'].dt.date
        grouped = df.groupby('date')
        rows = []
        for day, group in grouped:
            group = group.sort_values('timestamp').reset_index(drop=True)
            for ft in FIXED_TIMES:
                target_dt = datetime.combine(day, ft)
                now = datetime.now()
                if target_dt > now:
                    continue  # 未来时间点不记录
                closest_row = group.iloc[(group['timestamp'] - target_dt).abs().argsort().iloc[0]]
                row = {
                    "timestamp": target_dt,
                    f"{name}_price": closest_row['price'],
                    f"{name}_volume": closest_row['volume'],
                    f"{name}_abs_change": closest_row['abs_change']
                }
                rows.append(row)
        all_rows.append(pd.DataFrame(rows))

    if not all_rows:
        print("没有可更新的数据")
        return

    # 合并所有饰品数据
    combined_df = pd.DataFrame({"timestamp": sorted({row['timestamp'] for df in all_rows for _, row in df.iterrows()})})
    for df in all_rows:
        combined_df = pd.merge(combined_df, df, on="timestamp", how="left")

    combined_df.sort_values("timestamp", inplace=True)

    # ================== 用 openpyxl 写入数据，保留格式 ==================
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        if "wide" in wb.sheetnames:
            ws = wb["wide"]
        else:
            ws = wb.create_sheet("wide")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "wide"

    # 写表头（新表格才写）
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        for col, col_name in enumerate(combined_df.columns, 1):
            ws.cell(row=1, column=col, value=col_name)

    # 现有表格的数据行数
    existing_rows = ws.max_row

    # 找出新数据（比最后一行时间戳更新的部分）
    last_timestamp = None
    if existing_rows > 1:
        last_timestamp = ws.cell(row=existing_rows, column=1).value

    new_data = combined_df
    if last_timestamp:
        new_data = combined_df[combined_df["timestamp"] > pd.to_datetime(last_timestamp)]

    if new_data.empty:
        print("没有新的数据需要写入")
    else:
        for r_idx, row in enumerate(new_data.itertuples(index=False), existing_rows + 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(OUTPUT_FILE)
    print(f"{OUTPUT_FILE} 已更新，总行数: {ws.max_row}")


# ================== 主执行函数 ==================
def run_update():
    print(f"--- 数据刷新 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---")
    goods_data = {}
    for good in goods:
        df = fetch_price_history(good["good_id"])
        goods_data[good["name"]] = df
        if not df.empty:
            print(f"{good['name']} 最新价格: {df['price'].iloc[-1]:.2f} RMB")

    update_wide_excel(goods_data)
    print("程序执行完毕，已退出。")


if __name__ == "__main__":
    run_update()
